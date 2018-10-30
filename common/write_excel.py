import openpyxl
from openpyxl import load_workbook

def copy_excel(excelpath1,excelpath2):
    '''将一个excel表复制到另一个表'''
    wb2 =openpyxl.Workbook()
    wb2.save(excelpath2)

    '''读取数据：打开excel'''
    wb1=openpyxl.load_workbook(excelpath1)
    wb2=openpyxl.load_workbook(excelpath2)
    '''读取sheet'''
    sheets1=wb1.sheetnames
    sheets2=wb2.sheetnames
    '''选择sheet'''
    sheet1=wb1[sheets1[0]]
    sheet2=wb2[sheets2[0]]
    '''获取最大行数和列数'''
    max_row=sheet1.max_row
    max_column=sheet1.max_column

    for m in list(range(1,max_row+1)):
        '''
        ASCII 中 chr(97) = 'a'
        i:单元格编号
        cell1:获取单元格数据
        sheet2[i]:赋值到test单元格  
        '''
        for n in list(range(97,97+max_column)):
            n=chr(n)
            i='%s%d'%(n,m)
            cell1=sheet1[i].value
            sheet2[i].value=cell1
    wb2.save(excelpath2)
    wb1.close()
    wb2.close()


class Write_excel(object):
    def __init__(self,filename):
        '''打开文件并激活'''
        self.filename=filename
        self.wb=load_workbook(self.filename)
        self.ws=self.wb.active
    def write(self,row_n,col_n,value):
        '''在指定的行列键值队，写入数据 '''
        self.ws.cell(row_n,col_n).value=value
        self.wb.save(self.filename)

# if __name__=='__main__':
#     # copy_excel('debug_api.xlsx','test111.xlsx')
#     wt=Write_excel('test111.xlsx')
#     wt.write(4,5,'HELLEOP')
#     wt.write(4,6,'HELLEOP')